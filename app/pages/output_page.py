import streamlit as st
import pandas as pd
import os
from io import BytesIO
from datetime import datetime
import matplotlib.pyplot as plt
import matplotlib as mpl
import numpy as np
import re
from rule_based_insight import show_polymer_blend_insights
import plotly.express as px  # Put this at the top of your file
from rule_based_insight import plot_ingredient
from additional_details import display_polymer_info
import boto3
from dotenv import load_dotenv
import json
from io import StringIO

# Plot styling
mpl.rcParams.update({
    'axes.facecolor': '#1a1a1a',
    'axes.edgecolor': '#cccccc',
    'axes.labelcolor': '#ffffff',
    'xtick.color': '#cccccc',
    'ytick.color': '#cccccc',
    'figure.facecolor': '#1a1a1a',
    'text.color': '#ffffff',
    'legend.edgecolor': '#ffffff',
    'axes.titleweight': 'bold',
    'axes.titlesize': 14,
    'axes.labelsize': 12,
    'legend.frameon': True,
    'legend.facecolor': '#2a2a2a',
    'legend.labelspacing': 0.3
})


from openpyxl import load_workbook

def extract_all_links(filepath, sheet_name, column_name):
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]

    # Find the column index for the given column_name
    col_idx = None
    for idx, cell in enumerate(ws[1]):
        if cell.value == column_name:
            col_idx = idx + 1  # 1-based indexing in openpyxl
            break

    if col_idx is None:
        return {}

    link_map = {}

    for row in range(2, ws.max_row + 1):  # Skip header row
        cell = ws.cell(row=row, column=col_idx)

        # Case 1: Excel hyperlink object
        if cell.hyperlink:
            link_map[row - 2] = cell.hyperlink.target

        # Case 2: Cell contains plain URL text
        elif isinstance(cell.value, str) and cell.value.strip().startswith("http"):
            link_map[row - 2] = cell.value.strip()

        # Case 3: No link found
        else:
            link_map[row - 2] = ""

    return link_map


def process_value(val):
    if pd.isna(val):
        return np.nan
    val = str(val).strip().replace(',', '')
    val = re.sub(r'[*<>≥≤]', '', val)
    numbers = re.findall(r'\d+\.?\d*', val)
    if not numbers:
        return np.nan
    numbers = [float(n) for n in numbers]
    return round(sum(numbers) / len(numbers), 2)

def get_data(data_file):
    df = pd.read_excel(data_file, sheet_name="Clean Data")
    df_total = df
    keep_cols = ['Polymer Category', 'Polymer Grade', 'Type of Polymer','Supplier','Cost (USD/Kg)',
                 'Tensile Strength (MPa)','Elongation at break (%)','Estimated WVTR based on 100 µm thickness (conditions may vary)',
                 'Estimated WVTR based on 20 µm thickness (conditions may vary)',
                 'BBC (%)','Compostability','Polymer Grade_Link','Cost_Link','BBC_Link',
                 'Tensile Strength (MPa)_Link', 'Elongation at break (%)_Link']
    rename_dict = {
        'Cost (USD/Kg)': 'Cost',
        'Tensile Strength (MPa)': 'Tensile Strength',
        'Elongation at break (%)': 'Elongation at Break',
        'Estimated WVTR based on 100 µm thickness (conditions may vary)': 'WVTR',
        'Estimated WVTR based on 20 µm thickness (conditions may vary)':'WVTR-2',
        'BBC (%)':'BBC'
    }
    df = df[keep_cols].rename(columns=rename_dict)

    # Pull hyperlinks from Excel for Polymer Grade column
    # link_map = extract_all_links(data_file, "Clean Data", "Polymer Grade")
    # df["Polymer Grade_Link"] = df.index.map(link_map).fillna("")    

    df['Tensile Strength-n'] = df['Tensile Strength'].apply(process_value)
    df['Elongation at Break-n'] = df['Elongation at Break'].apply(process_value)
    df['WVTR'] = df['WVTR'].apply(process_value)
    df['WVTR-2'] = df['WVTR-2'].apply(process_value)

    df['BBC'] = df['BBC'].apply(process_value)
    return df,df_total



def show():
    st.title("📊 Recommended Polymers")
    filters = st.session_state.get("filters", {})
    data_file = "data/Workflow #1 - Data Needs.xlsx"
    df,df_total = get_data(data_file)

    # st.dataframe(df)

    def check(row):
        checks = {
            "Tensile Strength": filters['tensile'][0] <= row["Tensile Strength-n"] <= filters['tensile'][1],
            "Elongation at Break": filters['elongation'][0] <= row["Elongation at Break-n"] <= filters['elongation'][1],
            "WVTR": filters['wvtr'][0] <= row["WVTR"] <= filters['wvtr'][1],
            "WVTR-2": filters['wvtr'][0] <= row["WVTR-2"] <= filters['wvtr'][1],
            "Cost": filters['cost'][0] <= row["Cost"] <= filters['cost'][1],
            "BBC": filters['bbc'][0] <= row["BBC"] <= filters['bbc'][1]
        }

        # ✅ Certs: match against 'Compostability' column
        user_certs = filters.get('certs', [])
        compost_str = str(row.get("Compostability", ""))
        # st.write(user_certs)
        row_certs = [c.strip() for c in compost_str.split(",") if c.strip()]
        # st.write(row_certs)
        checks["Compostability"] = all(cert in row_certs for cert in user_certs)

        # ✅ Geo match
        if filters['geos']:
            checks["Continent"] = row.get("Continent", "") in filters["geos"]
        
        # st.write(checks)
        return checks

    results = []
    for _, row in df.iterrows():
        checks = check(row)
        passed = sum(checks.values())
        row_dict = row.to_dict()
        row_dict['Score'] = passed
        row_dict['Feedback'] = ""
        row_dict['Details'] = ""
        row_dict['Insights'] = ""

        for key, passed_check in checks.items():
            row_dict[f"{key}_Check"] = passed_check
        results.append(row_dict)

    ranked_df = pd.DataFrame(results)
    bio_df = ranked_df[ranked_df["Type of Polymer"] == "Biopolymer"].sort_values(by="Score", ascending=False)
    bench_df = ranked_df[ranked_df["Type of Polymer"] == "Benchmark"].sort_values(by="Score", ascending=False)
    combined_df = pd.concat([bio_df, bench_df]).reset_index(drop=True)

    all_columns = combined_df.columns.tolist()

    skip_columns = ["Score","Polymer Grade_Link","Cost_Link","BBC_Link","Tensile Strength (MPa)_Link","Elongation at break (%)_Link","Tensile Strength-n","Elongation at Break-n"]  # add any column name here
    main_columns = [col for col in all_columns if not col.endswith("_Check") and col not in skip_columns]

    #main_columns = [col for col in all_columns if not col.endswith("_Check")]

    st.subheader("🧪 Polymer Results Table")
    column_labels = {
    "Cost": "Cost ($/kg)",
    "BBC": "BBC (%)",
    "Tensile Strength": "Tensile Strength (MPa)",
    "Elongation at Break": "Elongation at Break (%)",
    "WVTR": "WVTR @100µm (g/m²·day)",
    "WVTR-2": "WVTR @20µm (g/m²·day)",
    "Polymer Grade": "Polymer Grade",
    "Polymer Category": "Polymer Category",
    "Type of Polymer": "Type of Polymer",
    "Supplier": "Supplier",
    "Feedback": "Feedback",
    "Details": "Details",
    "Insights": "Insights"
}



    header_cols = st.columns(len(main_columns))
    for i, col_name in enumerate(main_columns):
        label = column_labels.get(col_name, col_name)
        header_cols[i].markdown(
            f"""
            <div style="
                display: flex;
                flex-direction: column;
                justify-content: space-between;
                align-items: center;
                height: 70px;
                color: #eee;
                font-weight: 600;
                font-size: 0.95rem;
                text-align: center;
                padding: 0 4px;
            ">
                <div style="line-height: 1.2em;">{label}</div>
                <div style="width: 100%; border-top: 4px solid #8942E5; margin-top: 0.5rem;"></div>
            </div>
            """,
            unsafe_allow_html=True
        )

    # for i, col_name in enumerate(main_columns):
    #     label = column_labels.get(col_name, col_name)
    #     header_cols[i].markdown(
    #         f"""
    #         <div style='text-align:center; color:#eee; font-weight:600; font-size:0.95rem; padding-bottom:0.2rem;'>
    #             {label}
    #             <hr style='border: 2px solid #8942E5; margin: 0.4rem 0 0 0;' />
    #         </div>
    #         """,
    #         unsafe_allow_html=True
    #     )

    for idx, row in combined_df.iterrows():
        toggle_key = f"show_details_{idx}"
        insight_key = f"show_insights_{idx}"

        feedback_key = f"feedback_{idx}"
        button_key = f"button_{idx}"
        if toggle_key not in st.session_state:
            st.session_state[toggle_key] = False
        if insight_key not in st.session_state:
            st.session_state[insight_key] = False

        with st.container():
            st.markdown("<div style='margin-bottom: 16px;'>", unsafe_allow_html=True)
            display_cols = st.columns(len(main_columns))
            for i, col_name in enumerate(main_columns):
                bg_color = "#111" if col_name in ["Polymer Category", "Polymer Grade", "Score", "Type of Polymer","Supplier"] else (
                    "#1f3b1f" if row.get(f"{col_name}_Check") else "#3b1f1f"
                )
                text_color = "#eee"

                if col_name == "Feedback":

                    #popover for feedback
                    with display_cols[i].popover("Feedback"):
                        st.text_input("Feedback", key=feedback_key, label_visibility="collapsed", placeholder="Write here...")
                        if st.button("Submit Feedback", key=button_key+"submit"):
                            # save feedback to a text file and upload to s3 use .env for credentials if file does not exist create it if it does append to it
                            user = st.session_state.get("user", "anonymous")
                            ts = datetime.now().strftime("%Y%m%d-%H%M%S")
                            save_dir = f"Data/feedback/{user}/for_Polymers"
                            os.makedirs(save_dir, exist_ok=True)
                            file_txt = os.path.join(save_dir, f"feedback_for_{row['Polymer Category']}_{row['Polymer Grade']}_{idx}_{ts}.txt")

                            #upload to s3 if file exists read the file and add the new feedback to the end of the file
                            if os.path.exists(file_txt):
                                with open(file_txt, "r") as f:
                                    feedback = f.read()
                                feedback += f"\n{st.session_state[feedback_key]}"
                            else:
                                #create the file
                                with open(file_txt, "w") as f:
                                    f.write(st.session_state[feedback_key])
                                feedback = st.session_state[feedback_key]

                            #upload file_txt to s3
                            s3_client = boto3.client(
                                's3',
                                region_name=os.getenv('AWS_REGION'),
                                aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
                                aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY')
                            )
                            #upload to s3 in the user folder in the folder feedback_for_Polymers
                            bucket_name = "feedbackworkflow1"
                            object_key = f"{user}/feedback_for_Polymers/{row['Polymer Category']}_{row['Polymer Grade']}_{idx}_{ts}.txt"
                            s3_client.upload_file(
                                file_txt, 
                                bucket_name, 
                                object_key
                            )

                            #success message
                            st.success("Feedback submitted")

                        
                elif col_name == "Details":
                    if display_cols[i].button("Details", key=button_key):
                        st.session_state[toggle_key] = not st.session_state[toggle_key]
                elif col_name == "Insights":
                    if display_cols[i].button("Insights", key=f"insights_{idx}"):
                        st.session_state[insight_key] = not st.session_state[insight_key]

                elif col_name == "Polymer Grade":
                    grade = row[col_name]
                    link = row.get("Polymer Grade_Link", "")
                    if pd.notna(link) and isinstance(link, str) and link.startswith("http"):
                        display_cols[i].markdown(
                            f"""
                            <a href="{link}" target="_blank" style="
                                display: inline-block;
                                text-decoration: none;
                                background-color: {bg_color};
                                color: {text_color};
                                font-weight: bold;
                                padding: 6px 10px;
                                border-radius: 6px;
                                text-align: center;
                            ">{grade}</a>
                            """,
                            unsafe_allow_html=True
                        )
                    else:
                        display_cols[i].markdown(
                            f"""
                            <a href= "https://www.planeterthos.com/404" target="_blank" style="
                                display: inline-block;
                                text-decoration: none;
                                background-color: {bg_color};
                                color: {text_color};
                                font-weight: bold;
                                padding: 6px 10px;
                                border-radius: 6px;
                                text-align: center;
                            ">{grade}</a>
                            """,
                            unsafe_allow_html=True
                        )
                    # else:
                    #     display_cols[i].markdown(
                    #         f"<div style='border:1px solid #555; padding:10px; background-color:{bg_color}; color:{text_color}'>{grade}</div>",
                    #         unsafe_allow_html=True
                    #     )
                elif col_name == "Cost":
                    grade = row[col_name]
                    link = row.get("Cost_Link", "")
                    if pd.notna(link) and isinstance(link, str) and link.startswith("http"):
                        display_cols[i].markdown(
                            f"""
                            <a href="{link}" target="_blank" style="
                                display: inline-block;
                                text-decoration: none;
                                background-color: {bg_color};
                                color: {text_color};
                                font-weight: bold;
                                padding: 6px 10px;
                                border-radius: 6px;
                                text-align: center;
                            ">{grade}</a>
                            """,
                            unsafe_allow_html=True
                        )
                    else:
                        
                        display_cols[i].markdown(
                            f"""
                            <a href= "https://www.planeterthos.com/404" target="_blank" style="
                                display: inline-block;
                                text-decoration: none;
                                background-color: {bg_color};
                                color: {text_color};
                                font-weight: bold;
                                padding: 6px 10px;
                                border-radius: 6px;
                                text-align: center;
                            ">{grade}</a>
                            """,
                            unsafe_allow_html=True
                        )
                        # display_cols[i].markdown(
                        #     f"<div style='border:1px solid #555; padding:10px; background-color:{bg_color}; color:{text_color}'>{grade}</div>",
                        #     unsafe_allow_html=True
                        # )
                elif col_name == "BBC":
                    grade = row[col_name]
                    link = row.get("BBC_Link", "")
                    if pd.notna(link) and isinstance(link, str) and link.startswith("http"):
                        display_cols[i].markdown(
                            f"""
                            <a href="{link}" target="_blank" style="
                                display: inline-block;
                                text-decoration: none;
                                background-color: {bg_color};
                                color: {text_color};
                                font-weight: bold;
                                padding: 6px 10px;
                                border-radius: 6px;
                                text-align: center;
                            ">{grade}</a>
                            """,
                            unsafe_allow_html=True
                        )
                    else:
                        display_cols[i].markdown(
                            f"""
                            <a href= "https://www.planeterthos.com/404" target="_blank" style="
                                display: inline-block;
                                text-decoration: none;
                                background-color: {bg_color};
                                color: {text_color};
                                font-weight: bold;
                                padding: 6px 10px;
                                border-radius: 6px;
                                text-align: center;
                            ">{grade}</a>
                            """,
                            unsafe_allow_html=True
                        )
                        # display_cols[i].markdown(
                        #     f"<div style='border:1px solid #555; padding:10px; background-color:{bg_color}; color:{text_color}'>{grade}</div>",
                        #     unsafe_allow_html=True
                        # )
                elif col_name == "Tensile Strength":
                    grade = row[col_name]
                    link = row.get("Tensile Strength (MPa)_Link", "")
                    if pd.notna(link) and isinstance(link, str) and link.startswith("http"):
                        display_cols[i].markdown(
                            f"""
                            <a href="{link}" target="_blank" style="
                                display: inline-block;
                                text-decoration: none;
                                background-color: {bg_color};
                                color: {text_color};
                                font-weight: bold;
                                padding: 6px 10px;
                                border-radius: 6px;
                                text-align: center;
                            ">{grade}</a>
                            """,
                            unsafe_allow_html=True
                        )
                    else:
                        display_cols[i].markdown(
                            f"""
                            <a href= "https://www.planeterthos.com/404" target="_blank" style="
                                display: inline-block;
                                text-decoration: none;
                                background-color: {bg_color};
                                color: {text_color};
                                font-weight: bold;
                                padding: 6px 10px;
                                border-radius: 6px;
                                text-align: center;
                            ">{grade}</a>
                            """,
                            unsafe_allow_html=True
                        )
                        # display_cols[i].markdown(
                        #     f"<div style='border:1px solid #555; padding:10px; background-color:{bg_color}; color:{text_color}'>{grade}</div>",
                        #     unsafe_allow_html=True
                        # )

                elif col_name == "Elongation at Break":
                    grade = row[col_name]
                    link = row.get("Elongation at break (%)_Link", "")
                    if pd.notna(link) and isinstance(link, str) and link.startswith("http"):
                        display_cols[i].markdown(
                            f"""
                            <a href="{link}" target="_blank" style="
                                display: inline-block;
                                text-decoration: none;
                                background-color: {bg_color};
                                color: {text_color};
                                font-weight: bold;
                                padding: 6px 10px;
                                border-radius: 6px;
                                text-align: center;
                            ">{grade}</a>
                            """,
                            unsafe_allow_html=True
                        )
                    else:
                        display_cols[i].markdown(
                            f"""
                            <a href= "https://www.planeterthos.com/404" target="_blank" style="
                                display: inline-block;
                                text-decoration: none;
                                background-color: {bg_color};
                                color: {text_color};
                                font-weight: bold;
                                padding: 6px 10px;
                                border-radius: 6px;
                                text-align: center;
                            ">{grade}</a>
                            """,
                            unsafe_allow_html=True
                        )
                        # display_cols[i].markdown(
                        #     f"<div style='border:1px solid #555; padding:10px; background-color:{bg_color}; color:{text_color}'>{grade}</div>",
                        #     unsafe_allow_html=True
                        # )
                else:
                    display_cols[i].markdown(
                        f"<div style='border:1px solid #555; padding:10px; background-color:{bg_color}; color:{text_color}'>{row[col_name]}</div>",
                        unsafe_allow_html=True
                    )
            st.markdown("</div>", unsafe_allow_html=True)

        if st.session_state[toggle_key]:
            
            st.markdown(f"### Compostability data for {row['Polymer Category']}:", unsafe_allow_html=True)

            plot_ingredient("./data/Bio_Dis_Data.xlsx",row["Polymer Category"])

            display_polymer_info(df_total, row["Polymer Category"], row["Polymer Grade"])

            # st.markdown("### Additional details:", unsafe_allow_html=True)
            # st.markdown(html_table, unsafe_allow_html=True)  # ✅ Important!



        if st.session_state[insight_key]:
            st.markdown(f"### Insights for {row['Polymer Category']}")

            key = row["Polymer Category"] + row["Polymer Grade"]

            show_polymer_blend_insights(row["Polymer Category"],key)

            

        if idx + 1 == len(bio_df):
            st.markdown("<hr style='border: 1px dashed #444;'>", unsafe_allow_html=True)

    user = st.session_state.get("user", "anonymous")

    st.markdown("---")
    # ⭐️ Custom star rating widget
    def star_rating(label, key, default=0):
        # Initialize default value if not already present
        if key not in st.session_state:
            st.session_state[key] = default

        selected = st.session_state[key]

        st.markdown(f"**{label}**")

        # Render stars with correct fill state
        cols = st.columns(5)
        for i in range(5):
            star = "⭐️" if i < selected else "☆"
            if cols[i].button(star, key=f"{key}_btn_{i}"):
                st.session_state[key] = i + 1
                st.rerun()

        selected = st.session_state[key]
        st.caption(f"You selected: {selected} out of 5 ⭐")
        return selected

# 💬 Full feedback form inside popover
    # if st.button("💾 Save Table & Rate Results"):
    with st.popover("💾 Save Table &  💬  Rate Results"):
            st.markdown("**Thank you for using erthos' new tool!** Before you leave, please take a few minutes to share your thoughts.")

            st.markdown("**1️⃣ How likely are you to recommend this tool to others?**")
            st.markdown("_(1 = Not at all likely, 5 = Very likely)_")
            recommend_score = star_rating("Select your rating", key="recommend_score")

            st.markdown("---")
            improvement = st.text_area("2️⃣ What improvements would you like to see, to make this work better for you?", key="improvement")

            st.markdown("---")
            
            st.markdown("**3️⃣ Do you think this tool can help you develop your DOE or establish a hypothesis?**")
            st.markdown("_(1 = Not helpful, 5 = Extremely helpful)_")
            doe_rating = star_rating("Select your rating", key="doe_rating")

            doe_notes = st.text_area("Please provide any additional information (optional)", key="doe_notes")

            st.markdown("---")
            extra_feedback = st.text_area("4️⃣ Any other feedback for erthos? (optional)", key="extra_feedback")

            if st.button("✅ Submit Feedback"):
                user = st.session_state.get("user", "anonymous")
                ts = datetime.now().strftime("%Y%m%d-%H%M%S")
                save_dir = f"Data/Feedback/{user}"
                os.makedirs(save_dir, exist_ok=True)

                # Save the general feedback survey
                file_txt = os.path.join(save_dir, f"feedback_survey_{ts}.txt")
                with open(file_txt, "w") as f:
                    f.write(f"Recommendation Likelihood: {st.session_state.get('recommend_score', 0)}/5\n")
                    f.write(f"Improvement Suggestions: {improvement.strip()}\n")
                    f.write(f"DOE Relevance Score: {st.session_state.get('doe_rating', 0)}/5\n")
                    f.write(f"DOE Comments: {doe_notes.strip()}\n")
                    f.write(f"Other Feedback: {extra_feedback.strip()}\n")
                    
                    # Add a section for polymer-specific feedback
                    f.write("\n--- Polymer-Specific Feedback ---\n")
                    polymer_feedbacks = []
                    for idx, row in combined_df.iterrows():
                        feedback_key = f"feedback_{idx}"
                        feedback_text = st.session_state.get(feedback_key, "")
                        if feedback_text:  # Only write non-empty feedback
                            polymer_category = row['Polymer Category']
                            polymer_grade = row['Polymer Grade']
                            f.write(f"{polymer_category} - {polymer_grade}: {feedback_text}\n")
                            polymer_feedbacks.append({
                                "polymer_category": polymer_category,
                                "polymer_grade": polymer_grade,
                                "feedback": feedback_text
                            })

                # Update the DataFrame with the feedback before saving to Excel
                for idx, row in combined_df.iterrows():
                    feedback_key = f"feedback_{idx}"
                    combined_df.at[idx, "Feedback"] = st.session_state.get(feedback_key, "")
                
                # Save Excel with the feedback included
                results_file = os.path.join(save_dir, f"results_{ts}.xlsx")
                combined_df[main_columns].to_excel(results_file, index=False)

                # Save to S3 bucket
                try:
                    # Initialize S3 client
                    s3_client = boto3.client(
                        's3',
                        region_name=os.getenv('AWS_REGION'),
                        aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
                        aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY')
                    )
                    
                    # Define S3 bucket name and object keys
                    bucket_name = "feedbackworkflow1"
                    
                    # Upload Excel file to S3
                    excel_object_key = f"{user}/results_{ts}.xlsx"
                    s3_client.upload_file(
                        results_file, 
                        bucket_name, 
                        excel_object_key
                    )
                    
                    # Upload text feedback to S3
                    txt_object_key = f"{user}/feedback_survey_{ts}.txt"
                    s3_client.upload_file(
                        file_txt,
                        bucket_name,
                        txt_object_key
                    )
                    
                    # Upload JSON structured feedback for easier analysis
                    feedback_data = {
                        "user": user,
                        "timestamp": ts,
                        "survey": {
                            "recommendation_score": st.session_state.get('recommend_score', 0),
                            "improvement_suggestions": improvement.strip(),
                            "doe_relevance_score": st.session_state.get('doe_rating', 0),
                            "doe_comments": doe_notes.strip(),
                            "other_feedback": extra_feedback.strip()
                        },
                        "polymer_feedback": polymer_feedbacks
                    }
                    
                    json_object_key = f"{user}/feedback_{ts}.json"
                    json_data = StringIO(json.dumps(feedback_data, indent=2))
                    s3_client.put_object(
                        Bucket=bucket_name,
                        Key=json_object_key,
                        Body=json_data.getvalue()
                    )
                    
                    st.success("🎉 Thank you for your feedback! Data saved locally and to cloud storage.")
                    
                except Exception as e:
                    st.warning(f"Local feedback saved, but cloud upload encountered an issue: {str(e)}")
                    st.success("🎉 Thank you for your feedback and for being a part of the erthos journey.")
                
                st.info("Your input helps us grow and improve, so we can keep building tools that truly support you and your work.")

                # Offer download right after submission
                with open(results_file, "rb") as f:
                    st.download_button(
                        "📥 Download Your Results",
                        f.read(),
                        file_name=os.path.basename(results_file),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
    # if st.button("💾 Save Table & Rate Results"):
    # with st.popover("💾 Save Table &  💬  Rate Results"):
    #         st.markdown("**Thank you for using erthos' new tool!** Before you leave, please take a few minutes to share your thoughts.")

    #         recommend_score = st.slider("1️⃣ How likely are you to recommend this tool to others?", 1, 5, key="recommend_score")

    #         st.markdown("---")
    #         improvement = st.text_area("2️⃣ What improvements would you like to see, to make this work better for you?", key="improvement")

    #         st.markdown("---")
    #         doe_rating = st.slider("3️⃣ Do you think this tool can help you develop your DOE or establish a hypothesis?", 1, 5, key="doe_rating")
    #         doe_notes = st.text_area("Please provide any additional information (optional)", key="doe_notes")

    #         st.markdown("---")
    #         extra_feedback = st.text_area("4️⃣ Any other feedback for erthos? (optional)", key="extra_feedback")

    #         if st.button("✅ Submit Feedback"):
    #             user = st.session_state.get("user", "anonymous")
    #             ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    #             save_dir = f"Data/Feedback/{user}"
    #             os.makedirs(save_dir, exist_ok=True)

    #             file_txt = os.path.join(save_dir, f"feedback_survey_{ts}.txt")
    #             with open(file_txt, "w") as f:
    #                 f.write(f"Recommendation Likelihood: {recommend_score}/5\n")
    #                 f.write(f"Improvement Suggestions: {improvement.strip()}\n")
    #                 f.write(f"DOE Relevance Score: {doe_rating}/5\n")
    #                 f.write(f"DOE Comments: {doe_notes.strip()}\n")
    #                 f.write(f"Other Feedback: {extra_feedback.strip()}\n")

    #             st.success("🎉 Thank you for your feedback and for being a part of the erthos journey.")
    #             st.info("Your input helps us grow and improve, so we can keep building tools that truly support you and your work.")
        # with st.popover("⭐ Please rate the result and confirm export"):
        #     rating = st.slider("Rate the results (1 to 5 stars):", 1, 5, 4)
        #     ts = datetime.now().strftime("%Y%m%d-%H%M%S")
        #     save_dir = f"Data/Feedback/{user}"
        #     os.makedirs(save_dir, exist_ok=True)

        #     for idx in combined_df.index:
        #         combined_df.at[idx, "Feedback"] = st.session_state.get(f"feedback_{idx}", "")

        #     file_xlsx = f"{save_dir}/results_{ts}.xlsx"
        #     file_txt = f"{save_dir}/feedback_{ts}.txt"
        #     combined_df[main_columns].to_excel(file_xlsx, index=False)

        #     with open(file_txt, "w") as f:
        #         for _, row in combined_df.iterrows():
        #             f.write(f"{row['Polymer Category']}: {row['Feedback']}\n")
        #         f.write(f"\nOverall Rating: {rating} Stars\n")

        #     st.success(f"✅ Saved to {save_dir}")
        #     with open(file_xlsx, "rb") as f:
        #         st.download_button("📥 Download Excel Table", f.read(), file_name=os.path.basename(file_xlsx))

    st.markdown("---")
    st.subheader("📈 Analysis")


    with st.expander("1️⃣ Scatter Plot Explorer"):
        
        all_columns = df.columns.tolist()


        #skip_columns = ["Score","Polymer Grade_Link","Cost_Link","BBC_Link","Tensile Strength (MPa)_Link","Elongation at break (%)_Link","Tensile Strength-n","Elongation at Break-n"]  # add any column name here
        skip_columns = []
        main_columns = [col for col in all_columns if not col.endswith("_Check") and col not in skip_columns]
        
        x_col = st.selectbox("Select X Axis", main_columns, key="x_axis")
        y_col = st.selectbox("Select Y Axis", main_columns, key="y_axis")

        if st.button("Generate Scatter Plot"):
            fig = px.scatter(
                df,
                x=x_col,
                y=y_col,
                color="Type of Polymer",
                symbol="Type of Polymer",
                hover_data={"Polymer Category": True, "Polymer Grade": True},
                labels={x_col: x_col, y_col: y_col},
                title=f"{y_col} vs {x_col}",
                template="plotly_dark",
                height=550
            )

            fig.update_traces(marker=dict(size=12, opacity=0.85,
                                        line=dict(width=1, color='black')))
            fig.update_layout(
                title_font_size=18,
                legend=dict(bgcolor='rgba(0,0,0,0)', borderwidth=0),
                margin=dict(t=40, b=40, l=0, r=0),
            )

            st.plotly_chart(fig, use_container_width=True)

            # Optional: Export
            user = st.session_state.get("user", "anonymous")
            os.makedirs(f"Data/Feedback/{user}", exist_ok=True)
            buf = BytesIO()
            try:
                fig.write_image(buf, format="png")
                buf.seek(0)
                st.download_button("📥 Download Scatter Plot", buf, file_name=f"scatter_{x_col}_vs_{y_col}.png")
            except Exception as e:
                st.warning("Image export failed. Please install `kaleido` using `pip install -U kaleido`.")

            # Logging
            with open(f"Data/Feedback/{user}/analysis_log.txt", "a") as f:
                f.write(f"[{datetime.now()}] User {user} generated interactive scatter plot: {y_col} vs {x_col}\n")



    # with st.expander("1️⃣ Scatter Plot Explorer"):
    #     x_col = st.selectbox("Select X Axis", df.columns, key="x_axis")
    #     y_col = st.selectbox("Select Y Axis", df.columns, key="y_axis")
    #     if st.button("Generate Scatter Plot"):
    #         fig, ax = plt.subplots(figsize=(7, 5))

    #         for i, row in df.iterrows():
    #             x = row[x_col]
    #             y = row[y_col]
    #             label = str(row.get("Polymer Category", f"P{i}"))+str(row.get("Polymer Grade", f"P{i}"))
    #             ax.scatter(x, y, color='#4db8ff', edgecolor='#1a1a1a', alpha=0.8, s=80)
    #             ax.text(x, y, label, fontsize=8, color='white', alpha=0.7, ha='right')

    #         ax.set_xlabel(x_col)
    #         ax.set_ylabel(y_col)
    #         ax.set_title(f"{y_col} vs {x_col}")
    #         ax.grid(True, linestyle='--', alpha=0.3)
    #         st.pyplot(fig)

    #         os.makedirs(f"Data/Feedback/{user}", exist_ok=True)
    #         buf = BytesIO()
    #         fig.savefig(buf, format="png", bbox_inches='tight')
    #         buf.seek(0)
    #         st.download_button("📥 Download Scatter Plot", buf, file_name=f"scatter_{x_col}_vs_{y_col}.png")

    #         with open(f"Data/Feedback/{user}/analysis_log.txt", "a") as f:
    #             f.write(f"[{datetime.now()}] User {user} plotted {y_col} vs {x_col} with labeled scatter\n")

    with st.expander("2️⃣ Benchmark Comparison"):
        props = st.multiselect("Select properties", df.columns)

        # Combine Polymer Category + Grade as a unique label
        df["label"] = df["Polymer Category"] + " — " + df["Polymer Grade"]

        # Dropdowns for user to select full unique polymer label
        target_label = st.selectbox("Target polymer", df["label"].unique(), key="benchmark_target")
        benchmark_label = st.selectbox("Benchmark polymer", df["label"].unique(), key="benchmark_reference")

        if st.button("Generate Bar Chart"):
            try:
                # Extract category and grade for filtering
                target_cat, target_grade = target_label.split(" — ")
                benchmark_cat, benchmark_grade = benchmark_label.split(" — ")

                # Filter data for target and benchmark
                val_target = df[(df["Polymer Category"] == target_cat) & (df["Polymer Grade"] == target_grade)][props].iloc[0]
                val_benchmark = df[(df["Polymer Category"] == benchmark_cat) & (df["Polymer Grade"] == benchmark_grade)][props].iloc[0]

                # Create bar chart
                fig, ax = plt.subplots(figsize=(8, 5))
                x = range(len(props))
                ax.bar([i - 0.2 for i in x], val_target, width=0.4, label=target_label, color='#66c2a5')
                ax.bar([i + 0.2 for i in x], val_benchmark, width=0.4, label=benchmark_label, color='#fc8d62')
                ax.set_xticks(list(x))
                ax.set_xticklabels(props, rotation=45)
                ax.set_ylabel("Value")
                ax.set_title("Target vs Benchmark Comparison")
                ax.grid(axis='y', linestyle='--', alpha=0.3)
                ax.legend()
                st.pyplot(fig)

                # Save image and log
                os.makedirs(f"Data/Feedback/{user}", exist_ok=True)
                buf = BytesIO()
                fig.savefig(buf, format="png", bbox_inches='tight')
                buf.seek(0)
                st.download_button("📥 Download Bar Chart", buf, file_name=f"benchmark_{target_cat}_{target_grade}_vs_{benchmark_cat}_{benchmark_grade}.png")

                with open(f"Data/Feedback/{user}/analysis_log.txt", "a") as f:
                    f.write(f"[{datetime.now()}] User {user} compared {target_label} vs {benchmark_label} on {props}\n")

            except Exception as e:
                st.error("⚠️ Could not generate chart. Make sure both selections are valid and properties are selected.")

    # with st.expander("2️⃣ Benchmark Comparison"):
    #     props = st.multiselect("Select properties", df.columns)
    #     target = st.selectbox("Target polymer", df["Polymer Category"])
    #     benchmark = st.selectbox("Benchmark polymer", df["Polymer Category"])

    #     if st.button("Generate Bar Chart") and props:
    #         val_target = df[df["Polymer Category"] == target][props].iloc[0]
    #         val_benchmark = df[df["Polymer Category"] == benchmark][props].iloc[0]

    #         fig, ax = plt.subplots(figsize=(8, 5))
    #         x = range(len(props))
    #         ax.bar([i - 0.2 for i in x], val_target, width=0.4, label=target, color='#66c2a5')
    #         ax.bar([i + 0.2 for i in x], val_benchmark, width=0.4, label=benchmark, color='#fc8d62')
    #         ax.set_xticks(list(x))
    #         ax.set_xticklabels(props, rotation=45)
    #         ax.set_ylabel("Value")
    #         ax.set_title("Target vs Benchmark Comparison")
    #         ax.grid(axis='y', linestyle='--', alpha=0.3)
    #         ax.legend()
    #         st.pyplot(fig)

    #         os.makedirs(f"Data/Feedback/{user}", exist_ok=True)
    #         buf = BytesIO()
    #         fig.savefig(buf, format="png", bbox_inches='tight')
    #         buf.seek(0)
    #         st.download_button("📥 Download Bar Chart", buf, file_name=f"benchmark_{target}_vs_{benchmark}.png")

    #         with open(f"Data/Feedback/{user}/analysis_log.txt", "a") as f:
    #             f.write(f"[{datetime.now()}] User {user} compared {target} vs {benchmark} on {props}\n")

    if st.button("🔙 Back to Requirements"):
        st.session_state.page = "Input"
        st.rerun()
